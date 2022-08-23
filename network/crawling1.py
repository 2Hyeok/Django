# crawling1.py
# 네이버 환율 정보
# time을 주어 텀을 줌, 반복문을 계속 돌리면 ip를 막음

import urllib.request as req
from bs4 import BeautifulSoup as bs
# url="https://finance.naver.com/marketindex/exchangeList.naver"
# data = req.urlopen(url).read().decode("euc-kr") # 데이터를 받아옴, 네이버는 옛날인코딩 사용
# req.urlretrieve(url, "exchangeList.txt") # 파일을 저장함, 이전 파일저장방식과 동일함
#
# soup = bs(data, "html.parser")
# # 테그를 가져오기에는 너무 많기때문에 class 와 id 를 가져옴
# # tr 테그 밑의 클래스가 tit 밑의 a테그 혹은 tit밑의 a태그의 string 이런식으로 작성
# exchange = soup.select(".tit a") # class가 tit 인 a태그 여러개를 가져옴
#
# # for e in exchange : # 반복문 이용, 나라이름
# #     print(e.string.strip()) # 앞뒤 여백이 있기에 strip()을 이용해 지움
#
# sale = soup.select("td.sale") # class가 td의 sale인것
#
# for i in range(len(exchange)) : # exchange의 길이의 갯수만큼
#     print(exchange[i].string.strip(), "\t", sale[i].string.strip(), # 리스트, i번째방, 나라이름과 환율이 출력됨
#           sale[i].next_sibling.next_sibling.string.strip())
    
#-------------------------------------------------------------------------------------------------------

# 다음 뉴스정보
# url = "https://news.daum.net/"
# data = req.urlopen(url).read().decode("utf-8") # 데이터를 받아옴
# soup = bs(data, "html.parser")
#
# # 기사를 가져올것임
# import time # 텀을 주기위해 임포트
# titles = soup.select( "strong.tit_g a") # strong 태그 밑의 a 태그
# for i, title in enumerate(titles) :
#     if i > 4 : # 5개만 출력
#         break
#     print("<<", title.string.strip(), ">>") # 기사들의 제목 출력
#     # 기사내용은 주소가 필요함
#
#     article_url = title.attrs["href"] # href 속성을 꺼내라
#     # print(article_url) # 기사의 주소가 출력됨
#     article_data = req.urlopen(article_url).read().decode("utf-8")
#     # print(article_data) # 태그가 같이 있어 지저분함, 기사의 내용만 따로 스크랩핑 해주어야함
#     soup = bs(article_data, "html.parser") # 새로읽어온 기사를 다시 파싱함
#
#     # p 태그 안에 기사 내용이 있음, p태그를 묶고있는것을 받아오면됨 -> div의 harmonyContainer.section 의 p태그 라는 아이디
#
#     ps = soup.select("div#harmonyContainer section > p") # 자식선택자, section 바로 밑의 p 태그를 잡기위해 > 사용
#
#     for p in ps :
#         if p.string != None : # p의 string이 None 값이 아닌것만
#             print(p.string)
#
#     time.sleep(3) # 3초씩 멈추면서 출력, 기사까지 늘리기위해 초를 늘림
#     print()

#-----------------------------------------------------------------------------------------------------

# request 모듈
import requests
# # get 함수 사용
#
# req = requests.get("http://google.co.kr/robots.txt") # 텍스트 데이터를 얻어옴
# text = req.text # 내용만 읽어옴, setter, getter 가 아닌 post, get 방식의 get임
# print(text)
# # bin = req.content
# # print(bin) # 바이너리 형식이라 줄바꿈이 안됨
#
# # 이미지 얻어오기
# req = requests.get("https://post-phinf.pstatic.net/MjAyMTA3MTBfMTkz/MDAxNjI1OTAyMDg5MzE1.6DmgVNiDM_Ps7yu4MEBOLbd_ZMavurGObsUIexRdMN4g.XTVLJNNrExZZMzLKVs7SpED_PaTInYemWE77vyzbIZMg.JPEG/210710_%EC%B5%B8%EB%8B%A8_%EC%9D%B8%EC%8A%A4%ED%83%8010.jpg") # 이미지 주소가 필요함
# with open("test.jpg", "wb") as f: # 이미지이기에 쓰기의 바이너리로 작성
#     f.write(req.content) # 바이너리로 읽어 저장해라

# 한빛미디어 로그인
# url = "https://www.hanbit.co.kr/member/login_proc.php?m_id=filmal&m_passwd=abcd1234" # 이런식으로 링크를 줌, 아이디와 비밀번호를 바로 먹임
# 페이지 소스를 읽어 로그인 처리를함
# 개발자 도구
# https://www.hanbit.co.kr/member/login.php?id=aaa&passwd=111 이런식임

# url = "https://www.hanbit.co.kr/member/login_proc.php"
# # 처리하는 페이지 이름
# # 아이디 와 비밀번호를 어떤식으로 받는지
#
# # 링크를 바로주면 링크를 계속 바꿔주어야하기에 이방식으로 해줌
# login_info = {
#     "m_id" : "filmal",
#     "m_passwd" : "abcd1234"
#     }
#
# session = requests.session()
# con = session.post(url, login_info) # 세션의 값을 가져온다?
# con.raise_for_status() # 상태를 유지 시켜라
#
# url = "https://www.hanbit.co.kr/myhanbit/myhanbit.html" # 로그인된 마이페이지 주소
# data = session.get(url)
# # print(data) # 텍스트인지 바이너리인지 해주어야함
# data.raise_for_status()
# soup = bs(data.text, "html.parser") # 바이너리가 아닌 텍스트로 읽음
#
# mileage_title = soup.select_one("dl.mileage_section1 dt")
# mileage = soup.select_one("dl.mileage_section1 dd span").string # dl태그의 클래스가 mileage_section1 인 dd 의 sapn
# ecoin = soup.select_one("dl.mileage_section2 dd span").string
#
# print(mileage_title.string, ":", mileage.string, "점")
#
# ecoin_title = soup.select_one("dl.mileage_section2 dt")
# print(ecoin_title.string, ":", ecoin.string)

#  구매내역 출력
# select 로 한줄 가져온 후 반복문 돌림
# trs = soup.select("div.sm_myorder tr")
# for i in range(1, len(trs)) : # 두번째 tr부터 가라
#     print(trs[i].td.string, "\t", trs[i].a.string, "\t", trs[i].span.string) # tr의 td태그의 string, 날짜, 재목, 가격
    
#---------------------------------------------------------------------------------------------------------------

# selenium
# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service
# from webdriver_manager.chrome import ChromeDriverManager
# import time
#
# chrome_options = webdriver.ChromeOptions()
# driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), # 서비스의 대한 옵션 끝
#                           options=chrome_options)
#
# url = "http://naver.com"
# driver.get(url) # get 방식으로 호출
# time.sleep(5) # 5초간 멈추어라
#
# driver.save_screenshot("naver.png") # png 파일로 스크린샷 후 저장
# driver.quit() # 창을 닫아라

# selenium 이 제공해주는 함수를 써야함

#--------------------------------------------------------------------------------------------------------------------

# 구글 로그인
# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service
# from webdriver_manager.chrome import ChromeDriverManager
# import time
#
# from webdriver_manager.core.utils import ChromeType # 임포트 필요
# driver = webdriver.Chrome(service=Service(ChromeDriverManager(chrome_type=ChromeType.BRAVE).install()))
#
# url="https://accounts.google.com/signin/v2/identifier"
# driver.implicitly_wait(3) # 응답 후 3초기다림
# driver.get(url)
#
# # find_element로 통합됨
# driver.find_element("id", "identifierId").send_keys("instructor.set") # 아이디 입력
# idbutton = driver.find_element("id", "identifierNext")
# time.sleep(2)
# idbutton.click()
# time.sleep(5)
# 로그인 안됨, 막혀있음

#-----------------------------------------------------------------------------------------------------------------------

# 한빛 미디어 로그인 - selenium
# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service
# from webdriver_manager.chrome import ChromeDriverManager
# import time
#
# from webdriver_manager.core.utils import ChromeType # 임포트 필요
#
# url = "https://www.hanbit.co.kr/member/login.html"
# driver = webdriver.Chrome(service=Service(ChromeDriverManager(chrome_type=ChromeType.BRAVE).install()))
# driver.implicitly_wait(3)
# driver.get(url)
#
# # 여기 까지 페이지 뜨는지 확인
# driver.find_element("id", "m_id").send_keys("filmal")
# driver.find_element("id", "m_passwd").send_keys("abcd1234")
#
# loginbutton = driver.find_element("id", "login_btn")
# time.sleep(2)
# loginbutton.click()
# time.sleep(5)

#-----------------------------------------------------------------------------------------------------------------------

# 공공데이터 포털 - 국토교통부_아파트매매 실거래 상세 자료
import urllib.request as req
import urllib.parse as pa # url 키값만 따로 뺀 다른방법
url = "http://openapi.molit.go.kr/OpenAPI_ToolInstallPackage/service/rest/RTMSOBJSvc/getRTMSDataSvcAptTradeDev?LAWD_CD=41360&DEAL_YMD=202207&serviceKey=OUwSlot%2BeCADq1M9zzdj8Sh1Ni9C4Iiaj9VqSEnyvikodjynkoS1hrbUsP6mSENccvTJH%2FDe3s3y7i836Lk7ew%3D%3D" # 2015년 실제 아파트 매매가 한줄로 작성 해야함
# url = "http://openapi.molit.go.kr/OpenAPI_ToolInstallPackage/service/rest/RTMSOBJSvc/getRTMSDataSvcAptTradeDev"
# values = {
#     "LAWD_CD" : 41360,
#     "DEAL_YMD" : 202207,
#     "serviceKey" : "OUwSlot%2BeCADq1M9zzdj8Sh1Ni9C4Iiaj9VqSEnyvikodjynkoS1hrbUsP6mSENccvTJH%2FDe3s3y7i836Lk7ew%3D%3D"
#     }
#
# # params = pa.urlencode(values) # 인코딩된것을 또 인코딩 하면서 문제 발생, 서비스 키만 따로 빼서 인코딩
# url = url + "?" + params
# LAWD_CD=코드 앞자리 5자 &DEAL_YMD=년월 - 법정동 코드, 기간 바꾸고싶은 코드로 바꾸면됨

data = req.urlopen(url).read().decode("utf-8")
# print(data)
with open("apart202207.txt", "w", encoding="utf-8") as f :
    f.write(data)

# 한글을 영어로 바꿈
data = data.replace("거래금액", "price")
data = data.replace("법정동", "dong")
data = data.replace("아파트", "apartname")
data = data.replace("월", "month")
data = data.replace("일", "day")

soup = bs(data, "html.parser") # Beautiful soup으로 만들어줌
items = soup.select("item") # 태그 이름이 item인것을 찾아라
for item in items :
    # print(item)
    print(item.dong.string, item.month.string, "-", item.day.string, item.apartname.string, item.price.string)
    print()
    
#-----------------------------------------------------------------------------------------------------------------


# JSON 방식 분석(크롤링)
# JSON 데이터는 깃허브에서 분석용 자료 다운
# https://api.github.com/repositories

# import json # json 임포트 필요
# import os.path
# url = "https://api.github.com/repositories" # 깃허브 레퍼지토리 위치
# filename = "github.txt" # 파일의 이름
#
# if not os.path.exists(filename) :
#     req.urlretrieve(url, filename) # 파일로 저장
#
# # 읽을때는 json으로 읽어야함
# items = json.load(open(filename, "rt", encoding="utf-8")) # 파일을 열것, 텍스트파일을, utf-8방식으로
# # { {, , , , } } 이런방식의 txt문서임, json 이 알아서 잘라줌
# for item in items :
#     print(item["id"], item["name"], item["owner"]["login"]) # 키를 주어 꺼내면됨, json이 두개 묶여있을때 꺼내는 방법 item["owner"]["login"]

#---------------------------------------------------------------------------------------------------------------------------

# # CSV 데이터 분석
# import csv, codecs # 원래는 codes가 아닌 pandas를 이용해야하지만, 아직 배우지 않음
# filename = "test.csv" # 파일 이름 설정
# write = codecs.open(filename, "w", encoding="utf-8") # 파일 을 쓴다
# writer = csv.writer(write, delimiter=",") # delimiter는 구분자 라는뜻
# writer.writerow(["아이디","이름","가격"]) # 띄어쓰기 금지
# writer.writerow(["1000","HDD",200000])
# writer.writerow(["1001","SSD",300000])
# writer.writerow(["1002","Moniter",150000])
# writer.writerow(["1003","Mouse",20000])
# writer.writerow(["1004","Keyboard",30000]) # 그냥 잘랐기에 빈 인덱스가 들어감
# print("파일생성 완료")
#
# write.close() # 파일을 열어서 작성했기에 닫아주어야함
#
# readcsv = codecs.open(filename, "r", encoding="utf-8").read() # 열어서 바로 읽음
# # 콤마로 나누어져 있어서 잘라야함, pandas는 알아서 잘라줌
# data = []
# rows = readcsv.split("\r\n")
# # print(rows)
# for row in rows :
#     if row == "": # 빈문자열 일경우
#         break   # 멈춰라
#     cells = row.split(",") # 1차원 리스트
#     data.append(cells) # 2차원 리스트로 만들어줌
#
# for d in data :
#     print(d[0] + "\t" + d[1] + "\t" + d[2]) # 인덱스가

#---------------------------------------------------------------------------------------------------------------------------

# EXCEL 데이터 분석
# 라이브러리가 있어야함
# pip install openpyxl
import openpyxl
filename = "stat_100701.xlsx"
wb = openpyxl.load_workbook(filename) # 워크북으로 만듬
# ws = wb.worksheets[0] # 시트가 1번이면 1번, 2번이면 2번
ws = wb.active # 현제 열려있는 엑셀 페이지 ( 하나만 있을시 가능할듯 )
# for row in ws.rows : # 줄마다 데이터가 있기에 2중 for문 이용
#     for data in row :
#         if data.value == None : # 데이터의 값이 none 일경우 해당값을 출력값을 빈칸으로 두기
#             print("", end="\t")
#         else : # 아닐경우 출력
#             print(data.value, end="\t") # cell이 나옴, 그러기에 value를 붙임
#     print()

# 2차원 리스트로 만들기
# 9번과 10번만 출력
data = []
# 몇개만 빼서 할것
for row in ws.rows :
    if row[9] != None and row[10] != None :
        data.append([row[9].value, row[10].value])
del(data[0:4]) # 0 ~ 4까지 지움

# sort 해줄것
# 데이터가 달라 어떤데이터로 정렬할것인지 정해야함

data = sorted(data, key = lambda x : x[1], reverse=True) # 리턴함, 왼쪽것인지 오른쪽것인지 정렬을 정해줌, 기준이 되는 함수를 정해줌, 람다로 정해줌, x의 1번을 기준으로 정렬을한다 라는뜻
# 기본적으로 오름차순 정렬임, 내림차순이면 reverse 를 써줌
# 정렬이 이상함, 문자열로 인식을해 9가 제일 커서 90이 맨 위로옴
for d in data :
    print(d)
    
# 다시 파일로 저장을해줌
savefile="population.xlsx"
swb = openpyxl.Workbook() # 클래스 생성
sws = swb.active # 활성화 시켜줌
# sws = swb.create_sheet(title="인구") # 새로운 워크시트가 만들어짐, 주석, 기존시트에 만들것이기 떄문
for i, d in enumerate(data) : # 엑셀의 위치때문에 인덱스가 필요함
    sws.cell(row=i+1, column=1, value=d[0]) # 몇번째 행의 몇번째 열인지, 행열을 잡아줌, 단 엑셀은 행이 1번부터 있음 0번이 아님, 그래서 1을 더해줌
    sws.cell(row=i+1, column=2, value=d[1])
swb.save(savefile)
